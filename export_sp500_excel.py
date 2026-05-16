"""
export_sp500_excel.py
Genera un Excel con todas las empresas del S&P 500:
  - Nombre y ticker
  - Sector e industria (GICS)
  - Market cap (yfinance)
  - Peso en el índice (market cap / total market cap)
  - Dinero invertido aproximado basado en AUM del SPY (~$590B)
"""

import io
import requests
import pandas as pd
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP500_URL = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; research-script/1.0)"}

# AUM aproximado del ETF SPY (S&P 500 más grande) en USD — actualizar si es necesario
SPY_AUM_USD = 590_000_000_000  # ~$590B


SECTOR_COLORS = {
    "Information Technology":   "D6E4FF",
    "Communication Services":   "E2D6FF",
    "Health Care":               "D6FFE4",
    "Financials":                "FFF3D6",
    "Energy":                    "FFD6D6",
    "Materials":                 "D6FFF9",
    "Consumer Staples":          "FFF9D6",
    "Consumer Discretionary":    "FFE4D6",
    "Industrials":               "D6EEFF",
    "Utilities":                 "EED6FF",
    "Real Estate":               "D6FFD6",
}

SECTOR_ORDER = list(SECTOR_COLORS.keys())


def fetch_sp500() -> pd.DataFrame:
    print("Descargando lista del S&P 500 desde Wikipedia...")
    html = requests.get(SP500_URL, headers=HEADERS, timeout=15).text
    tables = pd.read_html(io.StringIO(html), attrs={"id": "constituents"})
    df = tables[0]
    df = df.rename(columns={
        "Symbol":               "Ticker",
        "Security":             "Empresa",
        "GICS Sector":          "Sector",
        "GICS Sub-Industry":    "Sub-Industria",
        "Headquarters Location":"Sede",
        "Date added":           "Fecha Añadida",
        "CIK":                  "CIK",
        "Founded":              "Fundada",
    })
    df["Ticker"] = df["Ticker"].str.replace(".", "-", regex=False)
    return df[["Ticker", "Empresa", "Sector", "Sub-Industria", "Sede"]]


def _get_market_cap(ticker: str) -> tuple[str, float | None]:
    try:
        v = yf.Ticker(ticker).fast_info.market_cap
        return ticker, float(v) if v else None
    except Exception:
        return ticker, None


def fetch_market_caps(tickers: list[str], workers: int = 20) -> dict[str, float | None]:
    result = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_get_market_cap, t): t for t in tickers}
        for i, future in enumerate(as_completed(futures), 1):
            ticker, cap = future.result()
            result[ticker] = cap
            print(f"  Progreso market cap: {i}/{len(tickers)}", end="\r")
    print()
    return result


def fmt_usd(v: float | None) -> str:
    if v is None:
        return "N/A"
    if v >= 1e12:
        return f"${v / 1e12:.2f}T"
    if v >= 1e9:
        return f"${v / 1e9:.2f}B"
    if v >= 1e6:
        return f"${v / 1e6:.1f}M"
    return f"${v:,.0f}"


def build_dataframe() -> pd.DataFrame:
    df = fetch_sp500()

    print("Obteniendo market caps desde yfinance...")
    caps = fetch_market_caps(df["Ticker"].tolist())
    df["Market Cap (USD)"] = df["Ticker"].map(caps)

    total = df["Market Cap (USD)"].sum()
    df["Peso en S&P 500 (%)"] = (df["Market Cap (USD)"] / total * 100).round(4)
    df["Dinero Invertido (SPY ~$590B)"] = (df["Peso en S&P 500 (%)"] / 100 * SPY_AUM_USD).round(0)

    df["Market Cap (fmt)"] = df["Market Cap (USD)"].apply(fmt_usd)
    df["Dinero Invertido (fmt)"] = df["Dinero Invertido (SPY ~$590B)"].apply(fmt_usd)

    sector_cat = pd.CategoricalDtype(categories=SECTOR_ORDER, ordered=True)
    df["Sector"] = df["Sector"].astype(sector_cat)
    df = df.sort_values(["Sector", "Market Cap (USD)"], ascending=[True, False], na_position="last")
    df["Sector"] = df["Sector"].astype(str)

    return df.reset_index(drop=True)


def style_worksheet(ws, df: pd.DataFrame):
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = {
        "A": 8,   # Ticker
        "B": 32,  # Empresa
        "C": 26,  # Sector
        "D": 38,  # Sub-Industria
        "E": 28,  # Sede
        "F": 18,  # Market Cap (fmt)
        "G": 14,  # Peso (%)
        "H": 26,  # Dinero Invertido (fmt)
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 36

    headers = ["Ticker", "Empresa", "Sector", "Sub-Industria", "Sede",
               "Market Cap", "Peso en S&P 500 (%)", "Dinero Invertido (SPY ~$590B)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.freeze_panes = "A2"

    prev_sector = None
    for r, (_, row) in enumerate(df.iterrows(), 2):
        sector = row["Sector"]
        color = SECTOR_COLORS.get(sector, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)

        values = [
            row["Ticker"],
            row["Empresa"],
            row["Sector"],
            row["Sub-Industria"],
            row["Sede"],
            row["Market Cap (fmt)"],
            row["Peso en S&P 500 (%)"],
            row["Dinero Invertido (fmt)"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if c in (1, 3):
                cell.alignment = Alignment(vertical="center", horizontal="center")
            if c == 7:
                cell.alignment = Alignment(vertical="center", horizontal="right")

        if sector != prev_sector and prev_sector is not None:
            for c in range(1, 9):
                existing = ws.cell(row=r, column=c).border
                ws.cell(row=r, column=c).border = Border(
                    left=thin, right=thin,
                    top=Side(style="medium", color="888888"),
                    bottom=thin,
                )
        prev_sector = sector


def build_summary_sheet(ws_sum, df: pd.DataFrame):
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    bold = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 18
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 30

    for c, h in enumerate(["Sector", "Empresas", "Market Cap Total", "Peso (%)", "Dinero Invertido (SPY)"], 1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    total_mcap = df["Market Cap (USD)"].sum()
    total_inv = df["Dinero Invertido (SPY ~$590B)"].sum()

    for r, sector in enumerate(SECTOR_ORDER, 2):
        grp = df[df["Sector"] == sector]
        if grp.empty:
            continue
        color = SECTOR_COLORS.get(sector, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        mcap = grp["Market Cap (USD)"].sum()
        peso = grp["Peso en S&P 500 (%)"].sum()
        inv = grp["Dinero Invertido (SPY ~$590B)"].sum()

        vals = [sector, len(grp), fmt_usd(mcap), f"{peso:.2f}%", fmt_usd(inv)]
        for c, val in enumerate(vals, 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center" if c != 1 else "left")

    last_r = len(SECTOR_ORDER) + 2
    ws_sum.cell(row=last_r, column=1, value="TOTAL S&P 500").font = bold
    ws_sum.cell(row=last_r, column=2, value=len(df)).font = bold
    ws_sum.cell(row=last_r, column=3, value=fmt_usd(total_mcap)).font = bold
    ws_sum.cell(row=last_r, column=4, value="100.00%").font = bold
    ws_sum.cell(row=last_r, column=5, value=fmt_usd(total_inv)).font = bold
    for c in range(1, 6):
        ws_sum.cell(row=last_r, column=c).border = border

    note_r = last_r + 2
    ws_sum.cell(row=note_r, column=1,
                value=f"* AUM SPY usado: {fmt_usd(SPY_AUM_USD)}  |  Datos: yfinance + Wikipedia  |  Fecha: {pd.Timestamp.today().date()}")
    ws_sum.cell(row=note_r, column=1).font = Font(italic=True, color="666666", size=9)
    ws_sum.merge_cells(f"A{note_r}:E{note_r}")


def main():
    df = build_dataframe()

    output = "/workspaces/Stocks/sp500_empresas.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: lista completa
        export_cols = ["Ticker", "Empresa", "Sector", "Sub-Industria", "Sede",
                       "Market Cap (fmt)", "Peso en S&P 500 (%)", "Dinero Invertido (fmt)"]
        df[export_cols].to_excel(writer, sheet_name="S&P 500 Empresas", index=False)
        ws = writer.sheets["S&P 500 Empresas"]
        style_worksheet(ws, df)

        # Sheet 2: resumen por sector
        writer.book.create_sheet("Resumen por Sector")
        ws_sum = writer.sheets["Resumen por Sector"]
        build_summary_sheet(ws_sum, df)

    print(f"\nExcel generado: {output}")
    print(f"Total empresas: {len(df)}")
    print(f"Total market cap: {fmt_usd(df['Market Cap (USD)'].sum())}")


if __name__ == "__main__":
    main()
